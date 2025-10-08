import pandas as pd
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QComboBox,
    QLineEdit, QHeaderView, QAbstractItemView, QMessageBox, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QIcon
from datetime import datetime, timedelta
from utils.utils import get_week_range, get_filename_wrt_date_filter_and_searchbox

class LogsWindow(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Attendance Logs")
        self.resize(800, 500)

        self.all_logs = []          # store all logs fetched from MongoDB
        self.filtered_logs = []     # store filtered logs based on search/filter

        main_layout = QVBoxLayout()

        # toop control bar layout
        top_layout = QHBoxLayout()

        # Date filter dropdown
        self.date_filter = QComboBox()
        self.date_filter.addItems(["Today", "This Week", "This Month"])
        self.date_filter.currentIndexChanged.connect(self.load_logs)
        top_layout.addWidget(self.date_filter)

        # Search box
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Search logs")
        self.search_box.textChanged.connect(self.apply_filters)
        top_layout.addWidget(self.search_box)

        # Reset Filter button
        self.btn_reset_filter = QPushButton("Reset filters")
        self.btn_reset_filter.clicked.connect(self.reset_filters)
        top_layout.addWidget(self.btn_reset_filter)

        # Excel export button
        self.btn_export_excel = QPushButton("Export")
        self.btn_export_excel.setIcon(QIcon("assets/icons/export-excel.png"))
        self.btn_export_excel.clicked.connect(self.export_filtered_logs)
        top_layout.addWidget(self.btn_export_excel)

        # Add top_layout into main layout
        main_layout.addLayout(top_layout)

        # Table to display logs
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Employee ID", "Name", "Department","Date", "Status", "Timestamp"])

        # Static UI styling (apply once)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setSortingEnabled(True)
        main_layout.addWidget(self.table)

        # Apply custom header and table styling
        self.setup_table_style()

        # Button to refresh logs
        self.btn_refresh = QPushButton("Refresh Logs")
        self.btn_refresh.clicked.connect(self.load_logs)
        main_layout.addWidget(self.btn_refresh)

        self.setLayout(main_layout)

        # Load logs on initialization
        self.load_logs()

    
    def setup_table_style(self):
        """Enhances the visual styling of the table and headers."""
        header = self.table.horizontalHeader()
        header.setStyleSheet("""
            QHeaderView::section {
                background-color: #f0f0f0;
                color: #000000;
                font-weight: bold;
                border-bottom: 1px solid #a0a0a0;
                padding: 2px;
            }
        """)

        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                selection-background-color: #dfe6e9;
                selection-color: #2d3436;
            }
            QTableWidget::item {
                padding: 2px;
            }
        """)


    def load_logs(self):
        filter_option = self.date_filter.currentText()
        today = datetime.today()

        if filter_option == "Today":
            start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        elif filter_option == "This Week":
            start_date = today - timedelta(days=today.weekday()) # Monday of this week
        elif filter_option == "This Month":
            start_date = today.replace(day=1)
        else:
            start_date = None

        query = {}
        if start_date:
            query["date"] = {"$gte": start_date.strftime("%Y-%m-%d")}

        self.all_logs = list(self.db.collection.find(query).sort("date", -1)) # Sort by the most recent
        self.apply_filters()    # show filtered (or all logs)


    def apply_filters(self):
        """Apply search-based filtering to logs"""
        search_text = self.search_box.text().lower()
        if not search_text:
            self.filtered_logs = self.all_logs
        else:
            self.filtered_logs = [
                log for log in self.all_logs
                if any(search_text in str(value).lower() for value in log.values())
            ]
        self.update_table()


    def reset_filters(self) -> None:
        """Reset search box, date filter, header sorting, and reload all logs."""
        self.btn_reset_filter.setEnabled(False)

        # Clear search and date filters
        self.search_box.clear()
        self.date_filter.setCurrentIndex(0)

        # Reset all column header sorts
        header = self.table.horizontalHeader()
        self.table.setSortingEnabled(False)
        header.setSortIndicatorShown(False)
        header.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)

        # Reload all logs
        self.load_logs()
        
        self.table.setSortingEnabled(True)  # Re-enable sorting
        header.setSortIndicatorShown(True)

        self.btn_reset_filter.setEnabled(True)


    def add_table_item(self, row, col, text):
        """Helper to create a centered, styled table item."""
        item = QTableWidgetItem(str(text))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        if col == 4:
            status_text = str(text).lower()
            if "present" in status_text:
                item.setForeground(Qt.GlobalColor.darkGreen)
            else:
                item.setForeground(Qt.GlobalColor.red)

        return item


    def export_filtered_logs(self):
        try:
            if not self.filtered_logs:
                QMessageBox.warning(self, "No Data", "There are no records to export.")
                return

            # Get filename based on date filter selection
            filter_option = self.date_filter.currentText()
            searchbox_text = self.search_box.text().lower()
            
            filename = get_filename_wrt_date_filter_and_searchbox(
                filter_option=filter_option, search_text=searchbox_text)
            
            # Open file dialog to select save location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Filtered logs",
                f"{filename}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return          # user cancelled
            
            # Convert logs to Dataframe and remove mongoDB _id
            df = pd.DataFrame(self.filtered_logs)
            if "_id" in df.columns:
                df = df.drop(columns=["_id"])

            # Ensure headers match the table headers in the ui
            headers = ["Employee ID", "Name", "Department", "Date", "Status", "Timestamp"]
            df = df.reindex(columns=["employee_id", "name", "department", "date", "status", "timestamp"])
            df.columns = headers

            # Export to Excel without index
            df.to_excel(file_path, index=False)

            # Excel styling
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = load_workbook(file_path)
            ws = wb.active
            ws.title = filename

            # define the border style for headers
            thick_border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )

            # Bold + center align header row
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thick_border

            # Center align all data cells and apply color coding to status cell
            status_col_idx = headers.index("Status") + 1    # 1-based index

            # define the border style once before the loop
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                    # if this cell is in the status column
                    if cell.column == status_col_idx:
                        status_text = str(cell.value).lower().strip()
                        if "present" in status_text:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # light green
                        elif "absent" in status_text:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # light red

            # Auto-adjust column width
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                adjusted_width = max_length + 2
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            # Save changes
            wb.save(file_path)

            QMessageBox.information(self, "Success", f"Logs exported successfully:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export logs:\n{str(e)}")


    def update_table(self):
        """Update QTableWidget with current filtered logs."""

        self.table.setSortingEnabled(False) # Temporarily disable sorting

        self.table.setRowCount(len(self.filtered_logs))

        for row, log in enumerate(self.filtered_logs):
            self.table.setItem(row, 0, self.add_table_item(row, 0, log.get("employee_id", "")))
            self.table.setItem(row, 1, self.add_table_item(row, 1, log.get("name", "")))
            self.table.setItem(row, 2, self.add_table_item(row, 2,log.get("department", "")))
            self.table.setItem(row, 3, self.add_table_item(row, 3,log.get("date", "")))
            self.table.setItem(row, 4, self.add_table_item(row, 4,log.get("status", "")))
            self.table.setItem(row, 5, self.add_table_item(row, 5,log.get("timestamp", "")))

        self.table.setSortingEnabled(True) # Re-enable sorting after population
   